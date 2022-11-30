from __future__ import print_function
import csv
from django.db.models.query import QuerySet
import datetime
import time
import os.path
from django.http import HttpResponse

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from base.models import Local, HorarioDeTrabajo, Producto
from BerSur.settings import TIME_ZONE, CALENDARS_IDS

from openpyxl import Workbook
from openpyxl.styles.fonts import Font

def obtener_numero_semana_del_mes(date_value):
    return (date_value.isocalendar()[1] - date_value.replace(day=1).isocalendar()[1] + 1)


def obtener_dias_de_semana(week, year):
    # as it starts with 0 and you want week to start from monday
    startdate = time.asctime(time.strptime(f'{year} %d 1' % week, '%Y %W %w'))
    startdate = datetime.datetime.strptime(startdate, '%a %b %d %H:%M:%S %Y')
    dates = [startdate.strftime('%d-%m-%Y')]
    for i in range(1, 6):
        day = startdate + datetime.timedelta(days=i)
        dates.append(day.strftime('%d-%m-%Y'))

    return dates


def mostrar_eventos(local: Local):
    """Shows basic usage of the Google Calendar API.
    Prints the start and name of the next 10 events on the user's calendar.
    """
    calendarId = CALENDARS_IDS[local.mail]
    # If modifying these scopes, delete the file token.json.
    SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('calendar', 'v3', credentials=creds)

        # Call the Calendar API
        now = datetime.datetime.utcnow().isoformat() + 'Z'  # 'Z' indicates UTC time
        print('Getting the upcoming 10 events')
        events_result = service.events().list(calendarId=calendarId, timeMin=now,
                                              maxResults=10, singleEvents=True,
                                              orderBy='startTime').execute()
        events = events_result.get('items', [])

        if not events:
            print('No upcoming events found.')
            return

        # Prints the start and name of the next 10 events
        for event in events:
            start = event['start'].get('dateTime', event['start'].get('date'))
            print(start, event['summary'])

    except HttpError as error:
        print('An error occurred: %s' % error)


def agregar_evento_al_calendario(horario, local):
    # Refer to the Python quickstart on how to setup the environment:
    # https://developers.google.com/calendar/quickstart/python
    # Change the scope to 'https://www.googleapis.com/auth/calendar' and delete any
    # stored credentials.
    SCOPES = ['https://www.googleapis.com/auth/calendar']

    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('calendar', 'v3', credentials=creds)

        calendarId = CALENDARS_IDS[local.mail]

        hora_parseada = horario.horario_ingreso.strftime("%Y-%m-%dT%H:%M:%S%Z")
        event = {
            'summary': 'Horario',
            'location': '800 Howard St., San Francisco, CA 94103',
            'description': 'a trabajar che',
            'start': {
                # 2015-05-28T09:00:00-07:00'

                'dateTime': hora_parseada,
                'timeZone': TIME_ZONE,
            },
            'end': {
                'dateTime': horario.horario_salida.strftime("%Y-%m-%dT%H:%M:%S%Z"),
                'timeZone': TIME_ZONE,
            },
            'recurrence': [
                'RRULE:FREQ=DAILY;COUNT=1'
            ],
            'attendees': [
                {'email': horario.empleado.usuario.mail},
            ],
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email', 'minutes': 24 * 60},
                    {'method': 'popup', 'minutes': 10},
                ],
            },
        }

        event = service.events().insert(calendarId=calendarId, body=event,
                                        sendNotifications=True).execute()
        print('Event created: %s' % (event.get('htmlLink')))

    except HttpError as error:
        print('An error occurred: %s' % error)


def agregar_eventos(local: Local):

    # [{empleado, horario_ingreso, horario_salida, local}, {empleado, horario_ingreso, horario_salida, local}, {empleado, horario_ingreso, horario_salida, local}]

    # horario_ingreso, horario_salida: {'local': 1, empleados[1,5]}
    # local: { horarios: [{'ingreso': 09:00, 'salida': 17:00, empleados: [1, 5]}, {'ingreso': 10:00, 'salida': 17:00, empleados: [6]}]}
    # local es redundante ?

    # horarios: {'ingreso': 09:00, 'salida': 17:00}
    # locales_ids = lista_horarios.values_list('local_id', flat=True).distinct()
    # locales = Local.objects.filter(id__in=locales_ids)

    # dame todos los horarios de esta semana de todos los empleados de este local
    hoy = datetime.date.today()
    semana_actual = hoy.isocalendar().week

    empleados_del_local = local.empleado.all()

    lista_horarios = HorarioDeTrabajo.objects.filter(
        empleado__in=empleados_del_local, horario_ingreso__week=semana_actual, horario_ingreso__year=hoy.year)

    horarios_agrupados = dict()
    for horario in lista_horarios:
        if (horario.horario_ingreso, horario.horario_salida) in horarios_agrupados:
            horarios_agrupados[(horario.horario_ingreso, horario.horario_salida)].append(
                horario.empleado)
        else:
            horarios_agrupados[(horario.horario_ingreso, horario.horario_salida)] = [
                horario.empleado]
    horarios_agrupados

    # Refer to the Python quickstart on how to setup the environment:
    # https://developers.google.com/calendar/quickstart/python
    # Change the scope to 'https://www.googleapis.com/auth/calendar' and delete any
    # stored credentials.
    SCOPES = ['https://www.googleapis.com/auth/calendar']

    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    try:
        service = build('calendar', 'v3', credentials=creds)

        calendarId = CALENDARS_IDS[local.mail]

        for horario, empleados in horarios_agrupados.items():
            listado_empleados = "\n".join(
                [f'{empleado.usuario.nombre} {empleado.usuario.apellido}' for empleado in empleados])
            event = {
                'summary': 'Horario',
                'location': local.direccion,
                'description': 'Horario designado para: \n' + listado_empleados,
                'start': {
                    'dateTime': horario[0].strftime("%Y-%m-%dT%H:%M:%S%Z"),
                    'timeZone': TIME_ZONE,
                },
                'end': {
                    'dateTime': horario[1].strftime("%Y-%m-%dT%H:%M:%S%Z"),
                    'timeZone': TIME_ZONE,
                },
                'recurrence': [
                    'RRULE:FREQ=DAILY;COUNT=1'
                ],
                'attendees': [
                    {'email': empleado.usuario.mail}
                    for empleado in empleados],
                'reminders': {
                    'useDefault': False,
                    'overrides': [
                        {'method': 'email', 'minutes': 24 * 60},
                        {'method': 'popup', 'minutes': 10},
                    ],
                },
            }

            event = service.events().insert(calendarId=calendarId, body=event,
                                            sendNotifications=True).execute()
            print('Event created: %s' % (event.get('htmlLink')))

    except HttpError as error:
        print('An error occurred: %s' % error)


def genereate_csv_response(labels, nombre_row, datos, nombre_archivo):

    response = HttpResponse(
        content_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{nombre_archivo}.csv"'},
    )
    writer = csv.writer(response)

    labels = [nombre_row] + labels
    writer.writerow(labels)
    #  Dict[str, Dict[int, List[int]]]
    for row_header, values in datos.items():
        # values: Dict[str, List[int] | int]
        row = [row_header] + values["montos"]
        writer.writerow(row)
        row.clear()

    return response


def genereate_productos_csv_response(productos: QuerySet[Producto]):
    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="productos.csv"'},
    )
    writer = csv.writer(response)

    writer.writerow(["Código", "Descripción", "Precio Costo", "Precio Venta", "Precio Mayoreo"])
    #  Dict[str, Dict[int, List[int]]]
    for producto in productos:
        # values: Dict[str, List[int] | int]
        row = [producto.codigo, producto.descripcion, f"${producto.precio_costo}",
               f"${producto.precio_venta}", f"${producto.precio_mayoreo}", producto.proveedor]
        writer.writerow(row)
        row.clear()

    return response

def export_productos_data(productos: QuerySet[Producto]):

    hoy = datetime.date.today()

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="productos - {hoy}.xlsx"'

    # create workbook
    wb = Workbook()
    sheet = wb.active

    # stylize header row
    # 'id','title', 'quantity','pub_date'
    c1 = sheet.cell(row = 1, column = 1) 
    c1.value = "Código"
    c1.font = Font(bold=True)

    c2 = sheet.cell(row= 1 , column = 2) 
    c2.value = "Descripción"
    c2.font = Font(bold=True)

    c3 = sheet.cell(row= 1 , column = 3) 
    c3.value = "Precio Costo"
    c3.font = Font(bold=True)

    c4 = sheet.cell(row= 1 , column = 4) 
    c4.value = "Precio Venta"
    c4.font = Font(bold=True)

    c5 = sheet.cell(row= 1 , column = 5) 
    c5.value = "Precio Mayoreo"
    c5.font = Font(bold=True)

    c6 = sheet.cell(row= 1 , column = 6) 
    c6.value = "Proveedor"
    c6.font = Font(bold=True)

    # export data to Excel
    rows = productos.values_list("codigo", "descripcion", "precio_costo", "precio_venta", "precio_mayoreo", "proveedor")
    for row_num, row in enumerate(rows, 1):
        # row is just a tuple
        for col_num, value in enumerate(row):
            c7 = sheet.cell(row=row_num+1, column=col_num+1) 
            c7.value = value

    wb.save(response)

    return response


def get_inicios_de_semana_unicos_locales_por_mes(locales: QuerySet[Local], mes, year):
    semanas_del_local = []
    for local in locales:
        for caja_chica in local.caja_chica_set.filter(semana__year=year, semana__month=mes):
            semana_del_local = caja_chica.semana.isocalendar().week
            semanas_del_local.append(semana_del_local)
    semanas_unicas_del_local = list(set(semanas_del_local))

    inicios_de_semana_unicos = []
    for semana in semanas_unicas_del_local:
        fecha = datetime.date.fromisocalendar(
            2022, semana, 1).strftime("%d/%m/%Y")
        inicios_de_semana_unicos.append(f'Semana del {fecha}')

    return inicios_de_semana_unicos
